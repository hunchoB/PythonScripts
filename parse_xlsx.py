import openpyxl

pathToFile = "D:\\vscodeProjects\\Python\\Software.xlsx"

wb_obj = openpyxl.load_workbook(pathToFile)
# sheet_obj = wb_obj.active

sheets = wb_obj.sheetnames

for i in sheets:
    ws_obj = wb_obj[i]
    print(ws_obj.title , ws_obj.max_row)


# max_col = sheet_obj.max_column

# for i in range(1, max_col + 1):
#     cell_obj = sheet_obj.cell(row = 1, column = i)
#     print(cell_obj.value)